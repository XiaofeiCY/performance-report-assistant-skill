#!/usr/bin/env python
"""Fill an Excel template copy while preserving workbook formatting.

Supported formats:
  .xlsx  — open and save with openpyxl (full support).
  .xlsm  — macro-enabled; requires keep_vba=True. Output must be .xlsm.
  .xltx  — template; copy to .xlsx output, warn that macros cannot be embedded.
  .xltm  — macro template; copy to .xlsm output. VBA may not carry over perfectly.
  .xls   — NOT supported; reject with a clear message.

Unsupported format rejection produces a readable error, not a stack trace.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any


SUPPORTED_TEMPLATE = {".xlsx", ".xlsm", ".xltx", ".xltm"}
MACRO_TEMPLATES = {".xlsm", ".xltm"}
TEMPLATE_ONLY = {".xltx", ".xltm"}


def flatten_mapping(raw: dict[str, Any]) -> dict[tuple[str, str], Any]:
    flattened: dict[tuple[str, str], Any] = {}
    for key, value in raw.items():
        if "!" in key:
            sheet, cell = key.split("!", 1)
            flattened[(sheet, cell)] = value
        elif isinstance(value, dict):
            for cell, cell_value in value.items():
                flattened[(key, cell)] = cell_value
        else:
            raise ValueError(f"Invalid mapping entry: {key}")
    return flattened


def merged_anchor(ws: Any, cell_ref: str) -> str:
    from openpyxl.cell.cell import MergedCell

    cell = ws[cell_ref]
    if not isinstance(cell, MergedCell):
        return cell_ref

    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return merged_range.start_cell.coordinate
    return cell_ref


def validate_suffixes(template_suffix: str, output_suffix: str) -> None:
    """Reject or warn about suffix mismatches before touching files."""
    if template_suffix not in SUPPORTED_TEMPLATE:
        raise ValueError(
            f"不支持模板文件格式 '{template_suffix}'。\n"
            f"支持的模板格式：{', '.join(sorted(SUPPORTED_TEMPLATE))}。\n"
            f"如果是 .xls 文件，请先在 Excel 中另存为 .xlsx 或 .xlsm 再使用。"
        )

    # .xlsm template must produce .xlsm output
    if template_suffix == ".xlsm" and output_suffix != ".xlsm":
        raise ValueError(
            f"模板是 .xlsm（含宏），但输出指定为 '{output_suffix}'。\n"
            f".xlsm 模板只能输出为 .xlsm，否则会丢失 VBA 宏。请将 --output 改为 .xlsm 后缀。"
        )

    # .xltm template must produce .xlsm output ONLY
    if template_suffix == ".xltm" and output_suffix != ".xlsm":
        raise ValueError(
            f"模板是 .xltm（含宏模板），但输出指定为 '{output_suffix}'。\n"
            f".xltm 模板只能输出为 .xlsm。请将 --output 改为 .xlsm 后缀。"
        )

    # .xltx template must produce .xlsx output ONLY
    if template_suffix == ".xltx" and output_suffix != ".xlsx":
        raise ValueError(
            f"模板是 .xltx（模板），但输出指定为 '{output_suffix}'。\n"
            f".xltx 模板只能输出为 .xlsx。请将 --output 改为 .xlsx 后缀。"
        )


def load_workbook_safe(path: Path) -> Any:
    """Load workbook with format-appropriate options.

    Macro-enabled files are loaded via a bytes buffer to avoid a
    Python 3.13 gc-time ZipFile warning about operations on closed files.
    """
    from openpyxl import load_workbook

    suffix = path.suffix.lower()
    if suffix in MACRO_TEMPLATES:
        try:
            # Read entire file into a bytes buffer so the zipfile inside
            # openpyxl does not reference the original file handle after close.
            import io
            data = path.read_bytes()
            return load_workbook(io.BytesIO(data), keep_vba=True)
        except Exception as e:
            raise RuntimeError(
                f"无法以保宏模式加载 {path}：{e}\n"
                f"含宏文件 (.xlsm/.xltm) 需要 openpyxl 以 keep_vba=True 加载。\n"
                f"如果当前 openpyxl 版本不支持保宏，请升级 openpyxl 或先将文件另存为 .xlsx。"
            ) from e
    return load_workbook(path)


def main() -> int:
    # Suppress "ValueError: I/O operation on closed file" during gc of
    # openpyxl-internal ZipFile objects. This is a benign Python-3.13
    # artifact; the underlying files are already correctly written.
    import sys as _sys

    _prev_hook = _sys.unraisablehook

    def _quiet_hook(args):
        if isinstance(args.exc_value, ValueError) and "closed file" in str(args.exc_value):
            return
        _prev_hook(args)

    _sys.unraisablehook = _quiet_hook

    parser = argparse.ArgumentParser(description="Fill values into a copied Excel template.")
    parser.add_argument("--template", required=True, help="Original template path (.xlsx/.xlsm/.xltx/.xltm).")
    parser.add_argument("--mapping", required=True, help="JSON mapping file.")
    parser.add_argument("--output", required=True, help="Output path.")
    args = parser.parse_args()

    from openpyxl.styles import Alignment

    template = Path(args.template)
    mapping_path = Path(args.mapping)
    output = Path(args.output)

    if not template.exists():
        raise FileNotFoundError(f"Template not found: {template}")
    if not mapping_path.exists():
        raise FileNotFoundError(f"Mapping file not found: {mapping_path}")

    template_suffix = template.suffix.lower()
    output_suffix = output.suffix.lower()

    # --- Input format guard ---
    if template_suffix == ".xls":
        print("错误：不支持 .xls 格式。")
        print("请先在 Excel 中打开该文件，另存为 .xlsx 或 .xlsm，然后用新文件作为模板。")
        print("方法：文件 → 另存为 → 选择 'Excel 工作簿 (*.xlsx)' 或 '启用宏的工作簿 (*.xlsm)'。")
        raise SystemExit(1)

    if output_suffix == ".xls":
        print("错误：不支持输出 .xls 格式。")
        print("请将 --output 改为 .xlsx 或 .xlsm 后缀。")
        raise SystemExit(1)

    # --- Suffix validation ---
    try:
        validate_suffixes(template_suffix, output_suffix)
    except ValueError as e:
        print(f"错误：{e}")
        raise SystemExit(1) from None

    output.parent.mkdir(parents=True, exist_ok=True)

    # Load template directly, modify in memory, save to output.
    # This avoids shutil.copyfile + load_workbook dual file access that
    # can cause "ValueError: I/O operation on closed file" on Windows.
    mapping = flatten_mapping(json.loads(mapping_path.read_text(encoding="utf-8")))
    wb = load_workbook_safe(template)

    if template_suffix in TEMPLATE_ONLY:
        print(f"注意：模板文件 ({template_suffix}) 已按输出格式 ({output_suffix}) 处理。")
        if template_suffix == ".xltm":
            print("提示：.xltm 宏模板中的 VBA 可能无法完整保留到输出文件，请人工检查。")

    for (sheet_name, cell_ref), value in mapping.items():
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet not found: {sheet_name}")
        ws = wb[sheet_name]
        target_ref = merged_anchor(ws, cell_ref)
        cell = ws[target_ref]
        cell.value = value
        cell.alignment = Alignment(
            horizontal=cell.alignment.horizontal,
            vertical=cell.alignment.vertical,
            text_rotation=cell.alignment.text_rotation,
            wrap_text=True,
            shrink_to_fit=cell.alignment.shrink_to_fit,
            indent=cell.alignment.indent,
        )

    wb.save(output)

    # Force garbage collection to trigger openpyxl's internal ZipFile __del__
    # while our custom unraisablehook is active. This suppresses the
    # "ValueError: I/O operation on closed file" noise from Python 3.13's
    # stricter file-handle checking during gc.
    import gc as _gc
    _gc.collect()

    if output_suffix == ".xlsm":
        print(f"已保存为 {output}（keep_vba=True，请验证宏是否完整）。")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
